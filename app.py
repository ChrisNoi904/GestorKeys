import os
import subprocess
import datetime
import time
import tempfile
import pathlib
import sys
import base64
import pymysql.cursors 
from xml.etree import ElementTree as ET

# Librerías necesarias para la migración y web
from flask import Flask, render_template, request, redirect, url_for, flash
from zeep import Client
from zeep.transports import Transport
from openpyxl import load_workbook
from zeep.helpers import serialize_object 

# =================================================================
#                         CONFIGURACIÓN Y GLOBALES
# =================================================================

# CUIT Emisor (se mantiene fijo o se gestiona como variable de entorno)
CUIT_EMISOR = 20216732716

# Servicios AFIP
WSAA_WSDL = "https://wsaa.afip.gov.ar/ws/services/LoginCms?wsdl"
PADRON_WSDL = "https://aws.afip.gov.ar/sr-padron/webservices/personaServiceA5?WSDL"
SERVICE_ID = "ws_sr_constancia_inscripcion"

# Archivo Excel de claves (solo para carga inicial si es necesario)
CLAVES_FILE = pathlib.Path(__file__).parent / "claves.xlsx"

# --- Configuración de Base de Datos (Desde Variables de Entorno de Render) ---
# Asegúrate de que estas variables de entorno estén configuradas en Render con tus credenciales de Hostinger
DB_HOST = os.environ.get("DB_HOST", "srv1591.hstgr.io") 
DB_USER = os.environ.get("DB_USER", "tu_usuario_hostinger") 
DB_PASSWORD = os.environ.get("DB_PASSWORD", "tu_contraseña_hostinger") 
DB_NAME = os.environ.get("DB_NAME", "u822656934_claves_cliente")
DB_PORT = int(os.environ.get("DB_PORT", 3306))

# --- Configuración del binario OpenSSL ---
OPENSSL_BIN = "openssl" 

# Inicializar Flask
app = Flask(__name__)
# Usar una clave secreta de entorno para producción
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "una_clave_secreta_fuerte") 

# =================================================================
#                         FUNCIONES DE BASE DE DATOS
# =================================================================

def get_db_connection():
    """Establece la conexión a la base de datos MySQL."""
    try:
        conn = pymysql.connect(
            host=DB_HOST,
            user=DB_USER,
            password=DB_PASSWORD,
            database=DB_NAME,
            port=DB_PORT,
            cursorclass=pymysql.cursors.DictCursor
        )
        return conn
    except Exception as e:
        app.logger.error(f"Error CRÍTICO al conectar a la base de datos: {e}")
        return None

# =================================================================
#                         LÓGICA AFIP (ADAPTADA A LA DB)
# =================================================================

def obtener_certificados_activos():
    """Recupera el contenido de los certificados activos de la DB."""
    conn = get_db_connection()
    if not conn: return None, None
    try:
        with conn.cursor() as cursor:
            # Selecciona el certificado activo y más reciente
            sql = "SELECT cert_content, key_content FROM certificados WHERE activo = 1 ORDER BY fecha_alta DESC LIMIT 1"
            cursor.execute(sql)
            result = cursor.fetchone()
            # Retorna el contenido de forma segura
            if result:
                return result.get('cert_content'), result.get('key_content')
            return None, None
    except Exception as e:
        app.logger.error(f"Error al obtener certificados: {e}")
        return None, None
    finally:
        if conn: conn.close()


def consultar_cuit_afip(cuit_consultado_str):
    """
    Función que maneja toda la lógica de conexión y consulta AFIP.
    Utiliza los contenidos de certificado/key directamente desde la DB.
    """
    
    cuit_consultado_str = cuit_consultado_str.strip()
    
    if not cuit_consultado_str.isdigit() or len(cuit_consultado_str) not in (11, 10):
        return None, "CUIT inválido."

    cert_content, key_content = obtener_certificados_activos()
    if not cert_content or not key_content:
        return None, "No se encontraron certificados AFIP activos en la base de datos."

    # Inicialización de rutas de archivos temporales para limpieza
    tmp_cert_path = None
    tmp_key_path = None
    tmp_lt_path = None
    tmp_cms_der_path = None

    try:
        # Creamos archivos temporales en el sistema de archivos de Render (/tmp)
        with tempfile.NamedTemporaryFile(delete=False) as tmp_cert, \
             tempfile.NamedTemporaryFile(delete=False) as tmp_key, \
             tempfile.NamedTemporaryFile(mode='w', delete=False, encoding="utf-8") as tmp_lt, \
             tempfile.NamedTemporaryFile(delete=False) as tmp_cms_der:

            # Escribir el contenido binario en los archivos temporales
            tmp_cert.write(cert_content)
            tmp_key.write(key_content)
            
            tmp_cert_path = tmp_cert.name
            tmp_key_path = tmp_key.name
            tmp_lt_path = tmp_lt.name
            tmp_cms_der_path = tmp_cms_der.name
            
            # Asegurarse de que se escriban los archivos antes de OpenSSL
            tmp_cert.close()
            tmp_key.close()

            # --- 1. CREAR LoginTicketRequest ---
            # FIX FINAL: Usar UTC, margen reducido y sufijo 'Z' para formato ISO 8601
            ahora_utc = datetime.datetime.utcnow() 
            
            # 1 minuto de margen hacia el pasado (GenerationTime).
            generation_time = (ahora_utc - datetime.timedelta(minutes=1)).strftime("%Y-%m-%dT%H:%M:%S") + "Z"
            # 10 minutos de margen hacia el futuro (ExpirationTime)
            expiration_time = (ahora_utc + datetime.timedelta(minutes=10)).strftime("%Y-%m-%dT%H:%M:%S") + "Z"

            login_ticket = f"""<?xml version="1.0" encoding="UTF-8"?>
<loginTicketRequest version="1.0">
  <header>
    <uniqueId>{int(time.time())}</uniqueId>
    <generationTime>{generation_time}</generationTime>
    <expirationTime>{expiration_time}</expirationTime>
  </header>
  <service>{SERVICE_ID}</service>
</loginTicketRequest>"""
            tmp_lt.write(login_ticket)
            tmp_lt.close() # Asegurarse de que el archivo se haya escrito

            try:
                # --- 2. FIRMAR EL TICKET (CMS) ---
                subprocess.check_call([
                    OPENSSL_BIN, 'smime', '-sign',
                    '-in', tmp_lt_path,
                    '-signer', tmp_cert_path,
                    '-inkey', tmp_key_path,
                    '-outform', 'DER',
                    '-nodetach', '-binary',
                    '-out', tmp_cms_der_path
                ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

                cms_der = pathlib.Path(tmp_cms_der_path).read_bytes()
                cms_b64 = subprocess.check_output([OPENSSL_BIN, 'base64'], input=cms_der).decode().strip()
                
                # --- 3. AUTENTICAR EN WSAA ---
                client_wsaa = Client(WSAA_WSDL, transport=Transport(timeout=20))
                ta_response = client_wsaa.service.loginCms(cms_b64)

                root_xml = ET.fromstring(ta_response) if isinstance(ta_response, str) else ET.fromstring(ET.tostring(ta_response, encoding="utf-8"))

                token = root_xml.findtext(".//token")
                sign = root_xml.findtext(".//sign")

                if not token or not sign:
                    return None, "Error de AFIP: No se pudo obtener token/sign del TA."
                
                # --- 4. CONSULTAR PADRÓN AFIP ---
                client_padron = Client(PADRON_WSDL, transport=Transport(timeout=20))
                respuesta = client_padron.service.getPersona_v2(token, sign, CUIT_EMISOR, int(cuit_consultado_str))

                # --- 5. PROCESAR RESPUESTA ---
                persona = getattr(respuesta, "personaReturn", respuesta)
                if persona is None:
                    return None, f"CUIT {cuit_consultado_str} no encontrado o sin datos."
                
                # Líneas de DIAGNÓSTICO (Opcional, se puede comentar para producción)
                # print("--- INICIO DIAGNÓSTICO AFIP (Zeep) ---")
                # print(serialize_object(persona))
                # print("--- FIN DIAGNÓSTICO AFIP (Zeep) ---")

                return persona, "Datos obtenidos (pendiente de formatear nombre en HTML)"

            except subprocess.CalledProcessError as e:
                app.logger.error(f"Error de OpenSSL al firmar: {e}")
                return None, f"Error de OpenSSL al firmar: {e}"
            except Exception as e:
                app.logger.error(f"Error en la consulta AFIP: {e}")
                return None, f"Error en la consulta AFIP: {e}"

    finally:
        # Eliminar archivos temporales de forma segura
        for path in [tmp_cert_path, tmp_key_path, tmp_lt_path, tmp_cms_der_path]:
            if path and os.path.exists(path):
                try:
                    os.unlink(path)
                except OSError as e:
                    app.logger.warning(f"No se pudo eliminar el archivo temporal {path}: {e}")

# =================================================================
#                         FUNCIONES CRUD/AUXILIARES
# =================================================================

def save_client_to_db(cuit, razon_social):
    """Guarda un cliente en la tabla clientes_afip si no existe."""
    conn = get_db_connection()
    if not conn: return False
    
    try:
        with conn.cursor() as cursor:
            # Verificar si existe
            cursor.execute("SELECT cuit FROM clientes_afip WHERE cuit = %s", (cuit,))
            if cursor.fetchone():
                return False # Ya existe

            # Insertar
            sql = "INSERT INTO clientes_afip (cuit, razon_social) VALUES (%s, %s)"
            cursor.execute(sql, (cuit, razon_social))
            conn.commit()
            return True
    except Exception as e:
        app.logger.error(f"Error al guardar cliente: {e}")
        return False
    finally:
        if conn: conn.close()


def load_data_from_excel():
    """Carga inicial de datos de clientes y claves desde archivos locales (solo para la migración inicial)."""
    if not CLAVES_FILE.exists():
        return False

    conn = None
    try:
        wb = load_workbook(CLAVES_FILE)
        ws = wb.active # Asume la primera hoja
        
        headers = [str(cell.value).lower().strip() if cell.value is not None else '' for cell in ws[1]]
        
        required_fields = ['cuit', 'descripcion_clave', 'usuario', 'clave']
        if not all(field in headers for field in required_fields):
             return False

        cuit_idx = headers.index('cuit')
        desc_idx = headers.index('descripcion_clave')
        user_idx = headers.index('usuario')
        clave_idx = headers.index('clave')
        
        # Intentar encontrar 'razon_social' si existe, sino usar un placeholder
        try:
             razon_social_idx = headers.index('razon_social')
        except ValueError:
             razon_social_idx = -1

        conn = get_db_connection()
        if not conn: return False

        with conn.cursor() as cursor:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > max(cuit_idx, desc_idx, user_idx, clave_idx):
                    cuit = str(row[cuit_idx]).strip().replace('-', '').replace(' ', '')
                    
                    if razon_social_idx != -1 and row[razon_social_idx]:
                        razon_social = str(row[razon_social_idx]).strip()
                    else:
                        razon_social = "Desconocida (Desde Excel)" 
                    
                    if cuit.isdigit() and len(cuit) >= 10:
                        # 1. Insertar Cliente (ignorar duplicados)
                        try:
                            cursor.execute("INSERT IGNORE INTO clientes_afip (cuit, razon_social) VALUES (%s, %s)", (cuit, razon_social))
                        except Exception:
                            pass # Ignoramos si ya existe

                        # 2. Insertar Clave
                        desc = str(row[desc_idx] or "").strip()
                        user = str(row[user_idx] or "").strip()
                        pw = str(row[clave_idx] or "").strip()
                        
                        sql = "INSERT INTO claves (cuit, descripcion_clave, usuario, clave) VALUES (%s, %s, %s, %s)"
                        cursor.execute(sql, (cuit, desc, user, pw))
            
            conn.commit()
            return True
    except Exception as e:
        app.logger.error(f"Error al cargar datos de Excel: {e}")
        return False
    finally:
        if conn: conn.close()


def format_afip_result_html(persona, cuit):
    """
    Formatea la respuesta del servicio AFIP a una cadena HTML legible,
    incluyendo Domicilio, Monotributo, Actividades, Impuestos y Regímenes.
    """
    if persona is None: return "No se recibieron datos de AFIP."
    
    html = f"<h2>Resultado de la Consulta AFIP (CUIT: {cuit})</h2>"
    
    dg = getattr(persona, 'datosGenerales', None)
    
    if not dg:
        return f"<p class='text-danger'>Error: No se obtuvieron datos generales para el CUIT {cuit}.</p>"

    # --- 1. CAPTURA DE NOMBRE/RAZÓN SOCIAL (manejo de FÍSICA/JURÍDICA) ---
    razon_social = getattr(dg, 'razonSocial', "")
    nombre = getattr(dg, 'nombre', "")
    apellido = getattr(dg, 'apellido', "")
    
    if razon_social:
        nombre_capturado = razon_social
    else:
        nombre_capturado = f"{nombre} {apellido}".strip()

    if not nombre_capturado:
        nombre_capturado = f"Nombre no encontrado (CUIT: {cuit})"
        
    # --- Datos Generales ---
    html += f"<h3>Datos Generales</h3>"
    html += f"<p><strong>Razón Social/Nombre:</strong> {nombre_capturado}</p>"
    html += f"<p><strong>Tipo Persona:</strong> {getattr(dg, 'tipoPersona', '—')}</p>"
    html += f"<p><strong>Estado Clave:</strong> {getattr(dg, 'estadoClave', '—')}</p>"
    
    # --------------------------- DOMICILIO ---------------------------
    dom = getattr(dg, 'domicilioFiscal', None)
    if dom:
        html += f"<h3>Domicilio Fiscal</h3>"
        html += f"<p><strong>Dirección:</strong> {getattr(dom, 'direccion', '—')}</p>"
        html += f"<p><strong>Localidad/Provincia:</strong> {getattr(dom, 'localidad', '—')} | {getattr(dom, 'descripcionProvincia', '—')}</p>"
        html += f"<p><strong>CP:</strong> {getattr(dom, 'codPostal', '—')}</p>"
        
    # =================================================================
    # DATOS DEL MONOTRIBUTO (Si existen)
    # =================================================================
    datos_monotributo = getattr(persona, 'datosMonotributo', None)
    
    if datos_monotributo:
        # 🟢 CORRECCIÓN: Intentar obtener la descripción de la categoría directamente del nodo principal
        desc_cat = getattr(datos_monotributo, 'descripcionCategoria', None)
        
        if not desc_cat:
            # Si no está en el nodo principal, buscar en el nodo anidado (para compatibilidad)
            categoria = getattr(datos_monotributo, 'categoriaMonotributo', None)
            if categoria:
                desc_cat = getattr(categoria, 'descripcionCategoriaMonotributo', '')
                
        if desc_cat:
            html += f"<h3>Datos del Monotributo</h3>"
            html += f"<p><strong>CATEGORÍA:</strong> {desc_cat}</p>"


    # =================================================================
    # DATOS DEL RÉGIMEN GENERAL (Incluye Jurídicas y Autónomos/Monotributistas)
    # =================================================================
    
    # Obtener el objeto que contiene las listas de detalle
    datos_regimen_general = getattr(persona, 'datosRegimenGeneral', None)
    
    # --------------------------- IMPUESTOS ---------------------------
    # Se extrae la lista de impuestos del objeto 'datosRegimenGeneral' (si existe)
    impuestos_list = getattr(datos_regimen_general, 'impuesto', []) if datos_regimen_general else []
    
    if impuestos_list:
        html += f"<h3>Impuestos (Inscripciones)</h3>"

        # Asegurar que sea una lista si Zeep devuelve un solo elemento sin envolver
        if not isinstance(impuestos_list, list):
            impuestos_list = [impuestos_list]

        for imp in impuestos_list:
            html += f"<p>- ID {getattr(imp, 'idImpuesto', '—')}: {getattr(imp, 'descripcionImpuesto', '—')}</p>"

    # --------------------------- ACTIVIDADES ---------------------------
    # Se extrae la lista de actividades del objeto 'datosRegimenGeneral' (si existe)
    actividades_list = getattr(datos_regimen_general, 'actividad', []) if datos_regimen_general else []
    
    if actividades_list:
        html += f"<h3>Actividades</h3>"
        
        if not isinstance(actividades_list, list):
            actividades_list = [actividades_list]

        for act in actividades_list:
            principal = ' (Principal)' if getattr(act, 'periodo', '') else ''
            html += f"<p>- Cód. {getattr(act, 'idActividad', '—')}: {getattr(act, 'descripcionActividad', '—')}{principal}</p>"

    # --------------------------- REGIMENES (RET/PER) ---------------------------
    # Se extrae la lista de regimenes del objeto 'datosRegimenGeneral' (si existe)
    regimenes_list = getattr(datos_regimen_general, 'regimen', []) if datos_regimen_general else []
    
    if regimenes_list:
        html += f"<h3>Otros Regímenes (Retenciones/Percepciones)</h3>"

        if not isinstance(regimenes_list, list):
            regimenes_list = [regimenes_list]

        for reg in regimenes_list:
            # La AFIP a veces envía 'idRegimen' y a veces 'id'
            reg_id = getattr(reg, 'idRegimen', None) or getattr(reg, 'id', '—')
            html += f"<p>- ID {reg_id}: {getattr(reg, 'descripcionRegimen', '—')}</p>"
            
    return html

# =================================================================
#                         RUTAS FLASK
# =================================================================

@app.route('/', methods=['GET', 'POST'])
def index():
    consulta_cuit = None
    razon_social = None
    error_message = None

    cert_content, _ = obtener_certificados_activos()
    if not cert_content:
        flash("Advertencia: No hay certificado AFIP activo. No se podrán realizar consultas al padrón.", 'warning')

    if request.method == 'POST':
        if 'cuit_consulta' in request.form:
            cuit_consulta = request.form['cuit_consulta']
            
            persona_data, mensaje_error = consultar_cuit_afip(cuit_consulta)
            
            if mensaje_error and persona_data is None:
                error_message = mensaje_error
            else:
                consulta_cuit = cuit_consulta
                
                # --- Lógica para obtener el nombre/razón social (para flash y DB) ---
                dg = getattr(persona_data, 'datosGenerales', None)
                if dg:
                    razon_social_temp = getattr(dg, 'razonSocial', "")
                    nombre_temp = getattr(dg, 'nombre', "")
                    apellido_temp = getattr(dg, 'apellido', "")
                    
                    if razon_social_temp:
                         razon_social = razon_social_temp
                    elif nombre_temp or apellido_temp:
                         razon_social = f"{nombre_temp} {apellido_temp}".strip()
                    else:
                         razon_social = None
                else:
                    razon_social = None

                if not dg or not razon_social:
                    error_message = mensaje_error or f"El CUIT {cuit_consulta} no devolvió datos generales o nombre/razón social."
                    
                    return render_template('index.html',
                                        cuit_consultado=consulta_cuit,
                                        error_message=error_message)

                # Pre-formatear el resultado para la web
                resultado_html = format_afip_result_html(persona_data, cuit_consulta)
                
                # Si se solicitó guardar el cliente
                if request.form.get('guardar_cliente'):
                    if save_client_to_db(consulta_cuit, razon_social):
                        flash(f"Cliente {razon_social} ({cuit_consulta}) guardado exitosamente.", 'success')
                    else:
                        flash(f"El cliente {cuit_consulta} ya estaba registrado.", 'info')
                
                return render_template('index.html', 
                                        resultado_html=resultado_html, 
                                        cuit_consultado=consulta_cuit,
                                        razon_social=razon_social)
        
        elif 'carga_inicial' in request.form:
            if load_data_from_excel():
                flash("Datos de clientes y claves migrados correctamente a la DB.", 'success')
            else:
                flash("Error al migrar datos de Excel. ¿Existe el archivo 'claves.xlsx'?", 'error')
            
            return redirect(url_for('index'))

    return render_template('index.html', error_message=error_message)

@app.route('/gestion_claves', methods=['GET', 'POST'])
@app.route('/gestion_claves/<cuit>', methods=['GET', 'POST'])
def gestion_claves(cuit=None):
    conn = get_db_connection()
    if not conn: 
        flash("Error crítico de conexión a la base de datos.", 'error')
        return redirect(url_for('index'))

    clientes = []
    claves = []
    
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT cuit, razon_social FROM clientes_afip ORDER BY razon_social")
            clientes = cursor.fetchall()
            
            if cuit:
                sql = "SELECT id, descripcion_clave, usuario, clave FROM claves WHERE cuit = %s"
                cursor.execute(sql, (cuit,))
                claves = cursor.fetchall()
            
            if request.method == 'POST':
                action = request.form.get('action')
                if action == 'add' and cuit:
                    desc = request.form['desc']
                    user = request.form['user']
                    pw = request.form['pw']
                    sql = "INSERT INTO claves (cuit, descripcion_clave, usuario, clave) VALUES (%s, %s, %s, %s)"
                    cursor.execute(sql, (cuit, desc, user, pw))
                    flash("Clave agregada.", 'success')
                
                elif action == 'update' and cuit:
                    key_id = request.form['key_id']
                    desc = request.form['desc']
                    user = request.form['user']
                    pw = request.form['pw']
                    sql = "UPDATE claves SET descripcion_clave=%s, usuario=%s, clave=%s WHERE id=%s AND cuit=%s"
                    cursor.execute(sql, (desc, user, pw, key_id, cuit))
                    flash("Clave actualizada.", 'success')

                elif action == 'delete' and cuit:
                    key_id = request.form['key_id']
                    sql = "DELETE FROM claves WHERE id=%s AND cuit=%s"
                    cursor.execute(sql, (key_id, cuit))
                    flash("Clave eliminada.", 'success')
                
                conn.commit()
                return redirect(url_for('gestion_claves', cuit=cuit))

    except Exception as e:
        flash(f"Error de DB en gestión de claves: {e}", 'error')
    finally:
        if conn: conn.close()

    return render_template('gestion_claves.html', clientes=clientes, claves=claves, cuit_seleccionado=cuit)

@app.route('/gestion_certificados', methods=['GET', 'POST'])
def gestion_certificados():
    conn = get_db_connection()
    if not conn: 
        flash("Error crítico de conexión a la base de datos.", 'error')
        return redirect(url_for('index'))

    tmp_cert_path = None 
    
    try:
        if request.method == 'POST':
            if 'cert_file' in request.files and 'key_file' in request.files:
                cert_file = request.files['cert_file']
                key_file = request.files['key_file']
                
                if cert_file and key_file:
                    cert_content = cert_file.read()
                    key_content = key_file.read()
                    
                    app.logger.info(f"--- INICIO POST CERTIFICADOS ---")
                    
                    try:
                        with tempfile.NamedTemporaryFile(delete=False) as tmp_cert:
                            tmp_cert.write(cert_content)
                            tmp_cert_path = tmp_cert.name
                        
                        vencimiento_output = subprocess.check_output([
                            OPENSSL_BIN, 'x509',
                            '-in', tmp_cert_path,
                            '-noout',
                            '-enddate'
                        ]).decode('utf-8').strip()

                        if not vencimiento_output.startswith("notAfter="):
                            raise ValueError(f"Formato de fecha de OpenSSL inesperado: {vencimiento_output}")
                        
                        vencimiento_str = vencimiento_output.split('=')[1].strip()
                        vencimiento = datetime.datetime.strptime(vencimiento_str, '%b %d %H:%M:%S %Y GMT').date()
                        app.logger.info(f"Vencimiento extraído por OpenSSL: {vencimiento}")
                        
                    except Exception as e:
                        app.logger.error(f"FALLO DE OPENSSL: Error al leer/validar certificado: {e}")
                        flash(f"Error al leer/validar el certificado (.crt). Asegúrese que sea un certificado X.509 válido en formato PEM y sin cifrar: {e}", 'error')
                        return redirect(url_for('gestion_certificados'))
                    finally:
                        if tmp_cert_path and os.path.exists(tmp_cert_path):
                            os.unlink(tmp_cert_path)
                            tmp_cert_path = None


                    with conn.cursor() as cursor:
                        cursor.execute("UPDATE certificados SET activo = 0")
                        sql = "INSERT INTO certificados (fecha_alta, fecha_vencimiento, nombre_archivo, cert_content, key_content, activo) VALUES (%s, %s, %s, %s, %s, 1)"
                        cursor.execute(sql, (datetime.datetime.now(), vencimiento, cert_file.filename, cert_content, key_content))
                        conn.commit()
                        flash(f"Certificado {cert_file.filename} cargado y activado. Vence el {vencimiento}.", 'success')
                    
                    return redirect(url_for('gestion_certificados'))

            elif 'delete_id' in request.form:
                cert_id = request.form['delete_id']
                with conn.cursor() as cursor:
                    cursor.execute("DELETE FROM certificados WHERE id = %s", (cert_id,))
                    conn.commit()
                    flash("Certificado eliminado del historial.", 'success')

        certificados = []
        today = datetime.date.today()
        
        try:
            with conn.cursor() as cursor:
                cursor.execute("SELECT id, fecha_alta, fecha_vencimiento, nombre_archivo, activo FROM certificados ORDER BY fecha_vencimiento DESC")
                raw_certificados = cursor.fetchall()
                
                for cert in raw_certificados:
                    if cert['fecha_alta']:
                        cert['fecha_alta'] = cert['fecha_alta'].strftime('%Y-%m-%d %H:%M:%S')
                    if cert['fecha_vencimiento']:
                        cert['fecha_vencimiento'] = cert['fecha_vencimiento'].strftime('%Y-%m-%d')
                    
                    certificados.append(cert)
            
        except Exception as e:
            app.logger.error(f"Error al obtener historial de certificados (GET): {e}")
            flash("Error al cargar el historial. Consulte los logs de Render para detalles.", 'error')
            
    except Exception as e:
        flash(f"Error general en la gestión de certificados: {e}", 'error')
    finally:
        if conn: conn.close()

    return render_template('gestion_certificados.html', certificados=certificados, today=today)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
